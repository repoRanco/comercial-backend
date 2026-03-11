"""
002-M-Sanyong.py
Transformación ETL - Marítimo - Sanyong
Genera un .xlsx con 2 hojas: Precios y Gastos
Uso: python 002-M-Sanyong.py <input.xlsx> <output.xlsx>
"""

import sys
import os
import logging
import pandas as pd
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============================================================================
# MAPEOS
# ============================================================================
MAPEO_COLUMNAS = {
    'Description': 'Marca', '品名': 'Marca', '品名 (Description)': 'Marca', '(Description)': 'Marca',
    'Variety': 'Variedad', '品种': 'Variedad', '品种 (Variety)': 'Variedad', '(Variety)': 'Variedad',
    'Size': 'Calibre', '规格': 'Calibre', '规格 (Size)': 'Calibre', '(Size)': 'Calibre',
    'Remarks': 'Envop', '备注': 'Envop', '备注 ( Remarks )': 'Envop', '( Remarks )': 'Envop',
    'Quantity': 'Cantidad', '数量': 'Cantidad', '数量 (Quantity)': 'Cantidad', '(Quantity)': 'Cantidad',
    'Average Price': 'PrecioUnitario', '单价': 'PrecioUnitario',
    '单价 (Average Price )': 'PrecioUnitario', '(Average Price )': 'PrecioUnitario',
    'Average Price ': 'PrecioUnitario'
}

MAPEO_COSTOS = {
    "增值税/Vat": (40, "01"),
    "客户佣金/Customer's Commission (6% of SELLING PRICE)": (40, "02"),
    "物流服务费/Forwarding charge": (40, "04"),
    "海关查验（中检查验装卸）/Fee for Custom clearance process": (40, "04"),
    "代理费/agent charge": (40, "04"),
    "单证费/Documents charge": (40, "04"),
    "入场费/Enter charge": (40, "05"),
    "场地费/Place rent": (40, "05"),
    "打冷费/cooling charge": (40, "06"),
    "冷处理费/Cold treatment": (40, "06"),
    "冷库费/Cold storage charge": (40, "06"),
    "装卸费/handling charges": (40, "07"),
    "叉车费/forklift truck charge": (40, "07"),
    "海关查验/Customs inspection": (40, "11"),
    "转运费/Change freight(SH-CZ)": (40, "10"),
    "转运费/Change freight": (40, "10"),
    "押车费/Rider stooped fee": (40, "10"),
    "海运费用/Ocean freight": (41, "01"),
    "空运费用/Air freight": (41, "01"),
}

SUBITEMS_40 = ["01","02","04","05","06","07","08","09","10","11"]
SUBITEMS_41 = ["01"]

PALABRAS_FIN_DATOS = ['合计/TOTAL SALES AMOUNT']
COLUMNAS_REQUERIDAS = ['Variedad', 'Calibre', 'Marca', 'Envop', 'Cantidad', 'PrecioUnitario']

ETIQUETA  = 'DL'
ESPECIE   = 'CE'
CATEGORIA = 'CAT1'

# ============================================================================
# UTILIDADES
# ============================================================================
def buscar_valor_numerico(row: pd.Series, desde_col: int) -> float:
    for i in range(desde_col, len(row)):
        try:
            valor = row.iloc[i]
            if pd.isna(valor):
                continue
            if isinstance(valor, (int, float)) and valor >= 0:
                return float(valor)
            s = str(valor).replace('￥','').replace(',','').replace(' ','').strip()
            if s and s != 'nan' and not any(c.isalpha() for c in s.replace('.','').replace('-','')):
                v = float(s)
                if v >= 0:
                    return v
        except Exception:
            continue
    return 0.0


def leer_excel_con_celdas_combinadas(ruta: str) -> pd.DataFrame:
    try:
        wb = load_workbook(ruta, data_only=True)
        ws = wb.active
        merged_ranges = {}
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            valor = ws.cell(min_row, min_col).value
            for col in range(min_col, max_col + 1):
                for row in range(min_row, max_row + 1):
                    merged_ranges[(row, col)] = valor
        max_col = ws.max_column
        fila_9, fila_10 = [], []
        for col_idx in range(1, max_col + 1):
            v9 = merged_ranges.get((9, col_idx), ws.cell(9, col_idx).value)
            v10 = merged_ranges.get((10, col_idx), ws.cell(10, col_idx).value)
            fila_9.append(v9 if v9 else "")
            fila_10.append(v10 if v10 else "")
        nombres_columnas = []
        for v9, v10 in zip(fila_9, fila_10):
            v9s = ' '.join(str(v9).replace('\n',' ').split()) if v9 and str(v9).strip() not in ["None",""] else ""
            v10s = ' '.join(str(v10).replace('\n',' ').split()) if v10 and str(v10).strip() not in ["None",""] else ""
            if v9s and v10s:
                nombre_col = f"{v10s} {v9s}"
            elif v9s:
                nombre_col = v9s
            elif v10s:
                nombre_col = v10s
            else:
                nombre_col = f"Columna_{len(nombres_columnas)}"
            nombres_columnas.append(nombre_col.strip())
        wb.close()
        df = pd.read_excel(ruta, header=None, skiprows=10)
        df.columns = nombres_columnas[:len(df.columns)]
        return df
    except Exception as e:
        logger.error(f"Error leyendo celdas combinadas: {e}")
        return pd.read_excel(ruta, header=9)


def limpiar_nombres_columnas(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = df.columns.astype(str)
    mapping = {}
    columnas_mapeadas = set()
    for col in df.columns:
        col_str = str(col).strip().replace('\n', ' ')
        if col_str in MAPEO_COLUMNAS:
            nombre_std = MAPEO_COLUMNAS[col_str]
            if nombre_std not in columnas_mapeadas:
                mapping[col] = nombre_std
                columnas_mapeadas.add(nombre_std)
                continue
        for patron, nombre_std in MAPEO_COLUMNAS.items():
            if (patron in col_str or col_str in patron) and nombre_std not in columnas_mapeadas:
                mapping[col] = nombre_std
                columnas_mapeadas.add(nombre_std)
                break
        if col not in mapping:
            col_lower = col_str.lower()
            if ('description' in col_lower or '品名' in col_str) and 'Marca' not in columnas_mapeadas:
                mapping[col] = 'Marca'; columnas_mapeadas.add('Marca')
            elif ('variety' in col_lower or '品种' in col_str) and 'Variedad' not in columnas_mapeadas:
                mapping[col] = 'Variedad'; columnas_mapeadas.add('Variedad')
            elif ('size' in col_lower or '规格' in col_str) and 'Calibre' not in columnas_mapeadas:
                mapping[col] = 'Calibre'; columnas_mapeadas.add('Calibre')
            elif ('remark' in col_lower or '备注' in col_str) and 'Envop' not in columnas_mapeadas:
                mapping[col] = 'Envop'; columnas_mapeadas.add('Envop')
            elif ('quantity' in col_lower or '数量' in col_str) and 'Cantidad' not in columnas_mapeadas:
                mapping[col] = 'Cantidad'; columnas_mapeadas.add('Cantidad')
            elif ('price' in col_lower or '单价' in col_str) and 'PrecioUnitario' not in columnas_mapeadas:
                mapping[col] = 'PrecioUnitario'; columnas_mapeadas.add('PrecioUnitario')
    df = df.rename(columns=mapping)
    faltantes = [c for c in COLUMNAS_REQUERIDAS if c not in df.columns]
    if faltantes:
        raise Exception(f"Columnas faltantes: {faltantes}")
    return df


def validar_fila(row: pd.Series) -> bool:
    marca = str(row.get('Marca', ''))
    if any(p in marca for p in PALABRAS_FIN_DATOS):
        return False
    for col in ['Variedad', 'Envop', 'Cantidad', 'PrecioUnitario']:
        if pd.isna(row.get(col)):
            return False
    try:
        if float(row['Cantidad']) <= 0:
            return False
        if float(row['PrecioUnitario']) < 0:
            return False
    except Exception:
        return False
    return True


# ============================================================================
# EXTRACCIÓN DE COSTOS
# ============================================================================
def obtener_costos(ruta: str) -> dict:
    df = pd.read_excel(ruta, header=None)
    resultado = {}
    procesados = set()
    for _, row in df.iterrows():
        for ci in range(len(row)):
            cell = str(row.iloc[ci]).strip() if pd.notna(row.iloc[ci]) else ""
            if len(cell) < 3:
                continue
            for texto, (item, subitem) in MAPEO_COSTOS.items():
                if texto == cell or texto in cell:
                    val = buscar_valor_numerico(row, ci + 1)
                    key = f"{texto}_{val}"
                    if val > 0 and key not in procesados:
                        k = (item, subitem)
                        resultado[k] = resultado.get(k, 0.0) + val
                        procesados.add(key)
                    break
    return resultado


# ============================================================================
# GENERACIÓN HOJAS
# ============================================================================
def generar_precios(filas: list) -> pd.DataFrame:
    df = pd.DataFrame(filas)
    df['_monto'] = df['Cantidad'] * df['PrecioUnitario']
    grp = df.groupby(['CodigoVariedad', 'CodigoCalibre', 'CodigoEnvop'], sort=False).agg(
        Cajas=('Cantidad', 'sum'), _monto_total=('_monto', 'sum')
    ).reset_index()
    grp['PrecioLiq'] = grp['_monto_total'] / grp['Cajas']
    rows = []
    for i, r in grp.iterrows():
        rows.append({
            'Fila': i+1, 'CodigEspecie': ESPECIE,
            'CodigoVariedad': r['CodigoVariedad'], 'CodigoEnvop': r['CodigoEnvop'],
            'CodigoCalibre': r['CodigoCalibre'], 'CodigoEtiqueta': ETIQUETA,
            'CodigoCategoria': CATEGORIA, 'Cajas': int(r['Cajas']),
            'PrecioLiq': round(r['PrecioLiq'], 7),
        })
    return pd.DataFrame(rows)


def generar_gastos(costos: dict, total_cajas: int) -> pd.DataFrame:
    rows = []
    for sub in SUBITEMS_40:
        val = costos.get((40, sub), 0.0)
        rows.append({'Especie': ESPECIE, 'CodigoItem': 40, 'CodigoSubItem': sub,
                     'Cajas': total_cajas, 'Valor': round(val, 2)})
    for sub in SUBITEMS_41:
        val = costos.get((41, sub), 0.0)
        rows.append({'Especie': ESPECIE, 'CodigoItem': 41, 'CodigoSubItem': sub,
                     'Cajas': total_cajas, 'Valor': round(val, 2)})
    return pd.DataFrame(rows)


# ============================================================================
# MAIN
# ============================================================================
def main():
    if len(sys.argv) < 3:
        print("Uso: python 002-M-Sanyong.py <input.xlsx> <output.xlsx>", file=sys.stderr)
        sys.exit(1)

    input_path  = sys.argv[1]
    output_path = sys.argv[2]

    if not os.path.exists(input_path):
        print(f"Archivo no encontrado: {input_path}", file=sys.stderr)
        sys.exit(1)

    df = leer_excel_con_celdas_combinadas(input_path)
    df = limpiar_nombres_columnas(df)

    filas = []
    for _, row in df.iterrows():
        if pd.isna(row.get('Envop')) or str(row.get('Envop')).strip() == "":
            row = row.copy(); row['Envop'] = 'Sin Envop'
        if validar_fila(row):
            filas.append({
                'CodigoVariedad': str(row['Variedad']).strip().upper(),
                'CodigoEnvop':    str(row['Envop']).strip().upper(),
                'CodigoCalibre':  str(row['Calibre']).strip().upper(),
                'Cantidad':       float(row['Cantidad']),
                'PrecioUnitario': float(row['PrecioUnitario']),
            })

    if not filas:
        print("ERROR: No se encontraron filas válidas", file=sys.stderr)
        sys.exit(1)

    total_cajas = int(sum(f['Cantidad'] for f in filas))
    costos = obtener_costos(input_path)
    df_precios = generar_precios(filas)
    df_gastos  = generar_gastos(costos, total_cajas)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_precios.to_excel(writer, sheet_name='Precios', index=False)
        df_gastos.to_excel(writer, sheet_name='Gastos',  index=False)

    print(f"FILAS:{len(df_precios)}")
    print(f"COLUMNAS:{len(df_precios.columns)}")


if __name__ == "__main__":
    main()
